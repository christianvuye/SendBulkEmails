import ezgmail
import mammoth
from openpyxl import load_workbook


# initialize ezgmail
# the first time this runs it will open a browser
ezgmail.init()

# print email adddress that is being used
print(ezgmail.EMAIL_ADDRESS)

# open excel file
print("Reading Excel file...")
wb = load_workbook(filename='Excel_Template.xlsx', read_only=True)
print("Excel file succesfully loaded.")\

print("Loading Excel sheet...")
ws = wb['Email Data']
print("Excel sheet loaded.")

# constant declarations and initialization
FIRST_ROW = 2
LAST_ROW = 1000000000
COLUMN_NAME = 1
COLUMN_EMAIL = 2
COLUMN_SUBJECT = 3
COLUMN_TEMPLATE = 4

# variable/iterator declaration and initialization
i = 0

# css to be added on top of HTML export
css = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <style>   
            * {
                font-family: "Gill Sans", "Gill Sans MT", Calibri, "Trebuchet MS",
                sans-serif;
            }
        
            table, th {
                border: 1px solid black;
                border-collapse: collapse;
            } 
            th {
                font-weight: normal;
                vertical-align: top;
                text-align: center;
                font-family: 'Gill Sans', 'Gill Sans MT', Calibri, 'Trebuchet MS', sans-serif;
            }
        </style>
    </head>
    <body>
    """
end_body_tag = "</body>"

# function declarations


def promp():
    print("Press Any Key To Exit.")
    input()
    exit()


def getCellValue(worksheet,  r, c):
    return worksheet.cell(row=r, column=c).value


def EndOfSheetReached():
    print("The end of the sheet has been reached.")
    printAmountofEmailsSent()


def printAmountofEmailsSent():
    print("DONE\t Total emails sent: " + str(i-FIRST_ROW))


# script logic
for i in range(FIRST_ROW, LAST_ROW):
    print(i)

    if getCellValue(ws, i, COLUMN_NAME) is not None and getCellValue(ws, i, COLUMN_EMAIL) is not None and getCellValue(ws, i, COLUMN_SUBJECT) is not None and getCellValue(ws, i, COLUMN_TEMPLATE) is not None:
        print(str(getCellValue(ws, i, COLUMN_NAME)).rstrip())
        doctor_name = str(getCellValue(ws, i, COLUMN_NAME)).rstrip()
        doctor_greeting_html = "<p>Dear " + doctor_name + ",</p>"

        print(str(getCellValue(ws, i, COLUMN_EMAIL)).rstrip())
        doctor_email = str(getCellValue(ws, i, COLUMN_EMAIL)).rstrip()

        print(str(getCellValue(ws, i, COLUMN_SUBJECT)).rstrip())
        subject = str(getCellValue(ws, i, COLUMN_SUBJECT)).rstrip()

        print(str(getCellValue(ws, i, COLUMN_TEMPLATE)).rstrip())
        template = str(getCellValue(ws, i, COLUMN_TEMPLATE)).rstrip()

        with open("FL_MZL Email Template.docx", "rb") as docx_file:
            result = mammoth.convert_to_html(
                docx_file, include_default_style_map=False)
            html_body = result.value
            messages = result.messages  # Any messages, such as warnings during conversion

        html_complete = css + doctor_greeting_html + html_body + end_body_tag

        ezgmail.send(doctor_email, subject, html_complete, mimeSubtype='html')

        print("Email sent to " + doctor_name + " with subject " + subject +
              " and email address " + doctor_email + " using template " + template)
    elif getCellValue(ws, i, COLUMN_NAME) is None:
        print("Row " + str(i) +
              " is missing data in the name column. Please fill in the missing name.")
        break
    elif getCellValue(ws, i, COLUMN_EMAIL) is None:
        print("Row " + str(i) +
              " is missing data in the email column. Please fill in the missing email.")
        break
    elif getCellValue(ws, i, COLUMN_SUBJECT) is None:
        print("Row " + str(i) +
              " is missing data in the subject column. Please fill in the missing subject.")
        break
    elif getCellValue(ws, i, COLUMN_TEMPLATE) is None:
        print("Row " + str(i) +
              " is missing data in the template column. Please fill in the missing template.")
        break
    else:
        break

EndOfSheetReached()
promp()
